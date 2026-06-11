import re
import openpyxl

def main():
    # 1. Parse _cityCoordinates from anagrafica_view.dart
    view_path = '/Users/giorgiomodoni/development/skyaudit/lib/features/anagrafica/anagrafica_view.dart'
    with open(view_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find the _cityCoordinates map
    map_match = re.search(r'static const Map<String, LatLng> _cityCoordinates = \{(.*?)\};', content, re.DOTALL)
    if not map_match:
        print("Could not find _cityCoordinates in anagrafica_view.dart")
        return

    map_content = map_match.group(1)
    # Extract keys
    keys = set()
    for line in map_content.split('\n'):
        line = line.strip()
        if not line or line.startswith('//'):
            continue
        key_match = re.match(r"['\"]([^'\"]+)['\"]\s*:", line)
        if key_match:
            keys.add(key_match.group(1).lower().strip())

    print(f"Loaded {len(keys)} city keys from anagrafica_view.dart.")

    # 2. Parse Excel file
    excel_path = '/Users/giorgiomodoni/development/skyaudit/data_mock/Database Domestic 8 aprile.xlsx'
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    if 'Anagrafica' not in wb.sheetnames:
        print("Anagrafica sheet not found in Excel")
        return

    sheet = wb['Anagrafica']
    excel_comuni = set()
    for r_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) > 41:
            val = row[41]
            if val is not None:
                val_str = str(val).strip()
                if val_str:
                    excel_comuni.add(val_str)

    print(f"Loaded {len(excel_comuni)} unique communes from Excel.")

    # 3. Find missing
    missing = []
    for c in sorted(excel_comuni):
        # We clean punctuation/spaces to check or check direct match
        c_clean = c.lower().strip()
        if c_clean not in keys:
            missing.append(c)

    print(f"Missing communes count: {len(missing)}")
    for m in missing:
        print(f"Missing: '{m}'")

if __name__ == '__main__':
    main()
